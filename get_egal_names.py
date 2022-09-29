import os
import requests
from openpyxl import load_workbook
import xlsxwriter

def findFiles():
    while True:
        try:
            m_path = os.getcwd()
            m_path = m_path.split('\\')[1:3]
            users = m_path[0]
            user = m_path[1]
            input_table = f"C:\{users}" + f"\{user}\путь до файла"
            output_table = f"C:\{users}" + f"\{user}\путь до файла"
            return input_table, output_table
        except IndexError:
            continue


def openInputTable(table):
    while True:
        try:
            wb = load_workbook(table)
            sheet = wb['1']
            maxRow = sheet.max_row
            return maxRow, sheet
        except IndexError:
            print("IndexError")
            continue

def getLink(row, sheet):
    link = sheet['A' + str(row)].value
    inn = sheet['B' + str(row)].value
    if "," in link:
        link = link[0:link.find(",")]
    return link, inn

def getCommercialName(name, headers):
    url = "https://google-search3.p.rapidapi.com/api/v1/search/q=" + name
    response = requests.request("GET", url, headers=headers).text
    start = response.find('title')+8
    end = response.find(',')
    response = response[start:end]
    response = response.replace('\\', "").replace('"', "")
    print(response)
    return response

def openOutputTable(output_table):
    wb = xlsxwriter.Workbook(output_table)
    ws = wb.add_worksheet()
    return wb, ws

def putInTable(commercialName, inn, row, ws):
    cell_name = 'A' + str(row)
    cell_inn = 'B' + str(row)
    ws.write(cell_name, commercialName)
    ws.write(cell_inn, inn)

def main():
    headers = {
        "X-User-Agent": "desktop",
        "X-Proxy-Location": "EU",
        "X-RapidAPI-Key": "key",
        "X-RapidAPI-Host": "google-search3.p.rapidapi.com"
    }
    for row in range(1, maxRow):
        link, inn = getLink(row, sheet)
        commercialName = getCommercialName(link, headers)
        putInTable(commercialName, inn, row, ws)
    wb.close()


if __name__ == "__main__":
    try:
        input_table, output_table = findFiles()
        maxRow, sheet = openInputTable(input_table)
        wb, ws = openOutputTable(output_table)
        main()
    except:
        wb.close()
        print("Ошибка!")
        # input()